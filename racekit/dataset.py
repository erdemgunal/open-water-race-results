from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from racekit.config import _provider_domain

RANK_ROLE_COLUMN = {
    "live": "gender_rank",
    "overall": "overall_rank",
    "gender": "gender_rank",
    "agegroup": "age_group_rank",
    "disabled": "disabled_rank",
}

SNAPSHOT_DIR = "snapshots"

COLUMN_ORDER = [
    "bib",
    "id",
    "first_name",
    "last_name",
    "full_name",
    "nation",
    "birth_year",
    "gender",
    "age_group",
    "gender_rank",
    "gender_rank_computed",
    "age_group_rank",
    "age_group_rank_computed",
    "overall_rank",
    "overall_rank_computed",
    "disabled_rank",
    "status",
    "time_text",
    "swim_seconds",
]

def to_nullable_int(s):
    return pd.to_numeric(s, errors="coerce").astype("Int64")

def add_derived_ranks(df):
    df = df.copy()
    fin = df["status"] == "FINISHED"

    df["overall_rank_computed"] = pd.Series([np.nan] * len(df), dtype="float64")
    df["gender_rank_computed"] = pd.Series([np.nan] * len(df), dtype="float64")
    df["age_group_rank_computed"] = pd.Series([np.nan] * len(df), dtype="float64")

    if fin.any():
        df.loc[fin, "overall_rank_computed"] = df.loc[fin, "swim_seconds"].rank(method="min")
        df.loc[fin, "gender_rank_computed"] = (
            df.loc[fin].groupby("gender")["swim_seconds"].rank(method="min")
        )
        df.loc[fin, "age_group_rank_computed"] = (
            df.loc[fin].groupby("age_group")["swim_seconds"].rank(method="min")
        )

    for c in ("overall_rank_computed", "gender_rank_computed", "age_group_rank_computed"):
        df[c] = to_nullable_int(df[c])

    for official, computed in (
        ("overall_rank", "overall_rank_computed"),
        ("gender_rank", "gender_rank_computed"),
        ("age_group_rank", "age_group_rank_computed"),
    ):
        if official not in df.columns:
            df[official] = pd.NA
        if df[official].isna().all():
            df[official] = df[computed]
    return df

def build_dataset(cfg, frames):
    base = frames["participants"].copy()
    if base.empty:
        raise RuntimeError("participants listesi boş veri seti üretilemedi.")
    for c in ("status", "time_text", "swim_seconds",
              "overall_rank", "gender_rank", "age_group_rank", "disabled_rank"):
        if c in base.columns:
            base = base.drop(columns=[c])

    res = frames[cfg.result_role].copy()
    res_cols = ["bib", "status", "time_text", "swim_seconds"]
    rank_col = RANK_ROLE_COLUMN.get(cfg.result_role)
    if rank_col and rank_col in res.columns:
        res_cols.append(rank_col)
    merged = base.merge(res[res_cols].drop_duplicates("bib"), on="bib", how="left")
    merged["status"] = merged["status"].fillna("DNS")
    merged["time_text"] = merged["time_text"].where(merged["time_text"].notna(), None)
    merged["swim_seconds"] = to_nullable_int(merged["swim_seconds"])

    for role, frame in frames.items():
        if role == "participants" or role == cfg.result_role:
            continue
        col = RANK_ROLE_COLUMN.get(role)
        if col is None or frame.empty or col not in frame.columns:
            continue
        sub = frame[["bib", col]].drop_duplicates("bib").copy()
        merged = merged.merge(sub, on="bib", how="left", suffixes=("", "_dup"))
        dup = col + "_dup"
        if dup in merged.columns:
            merged[col] = merged[col].fillna(merged[dup])
            merged = merged.drop(columns=[dup])

    merged = add_derived_ranks(merged)

    if "full_name" not in merged.columns:
        merged["full_name"] = pd.NA
    missing_fn = merged["full_name"].isna()
    if missing_fn.any() and "first_name" in merged.columns:
        fn = merged["first_name"].astype("string").fillna("").str.strip()
        ln = merged["last_name"].astype("string").fillna("").str.strip()
        merged.loc[missing_fn, "full_name"] = (fn[missing_fn] + " " + ln[missing_fn]).str.strip()
        merged.loc[(missing_fn) & (merged["full_name"] == ""), "full_name"] = pd.NA

    for c in COLUMN_ORDER:
        if c not in merged.columns:
            merged[c] = pd.NA

    merged = merged[COLUMN_ORDER].copy()
    merged = merged.sort_values(
        ["overall_rank_computed", "bib"], na_position="last"
    ).reset_index(drop=True)
    return merged

COLUMN_DOCS = {
    "bib": ("Start numarası (yarışçı numarası)", "int (Int64)", "participants + sonuç", "Birleştirme anahtarı"),
    "id": ("RaceResult internal katılımcı ID", "int (Int64)", "participants + sonuç", "sağlayıcı üye kaydı"),
    "first_name": ("Ad", "str", "participants", "Yalnızca sağlayıcı ayrık veriyorsa doldurulur (İstanbul) birleşik isim veren kaynaklarda boş kalır"),
    "last_name": ("Soyad", "str", "participants", "Yalnızca sağlayıcı ayrık veriyorsa doldurulur (İstanbul) birleşik isim veren kaynaklarda boş kalır"),
    "full_name": ("Ad Soyad (tam)", "str", "participants / sonuç", "Çanakkale FLNAME İstanbul'da first+last birleşimi"),
    "nation": ("Ülke kodu (ISO 3166-1 alpha-2)", "str", "participants + sonuç", "NATION.FLAG img path'inden çıkarılır"),
    "birth_year": ("Doğum yılı", "int (Int64)", "participants + sonuç", "Yaş = yıl - birth_year"),
    "gender": ("Cinsiyet: F / M", "str", "participants + sonuç", "GenderMF, GenderFix ya da grup anahtarı"),
    "age_group": ("Yaş grubu kategorisi", "str", "participants", "İstanbul: 'Male B(19-24)' Çanakkale: 'Male 19-24'"),
    "gender_rank": ("Resmi cinsiyet sıralaması", "int (Int64)", "live (İst.) / GenderRankList (Çnk.)", "Eşit görünen sürelerde gizli alt-saniye ile kırılır"),
    "gender_rank_computed": ("Cinsiyet sıralaması (hesaplanan)", "int (Int64)", "türetilmiş", "rank(method='min') berabere süreler aynı sırayı paylaşır"),
    "age_group_rank": ("Yaş grubu sıralaması", "int (Int64)", "AgeGroupRankList (Çnk.) / türetilmiş", "İstanbul'da resmi yayın yok hesaplanır"),
    "age_group_rank_computed": ("Yaş grubu sıralaması (hesaplanan)", "int (Int64)", "türetilmiş", "rank(method='min')"),
    "overall_rank": ("Genel sıralama", "int (Int64)", "OverallRankList (Çnk.) / türetilmiş", "İstanbul'da resmi genel sıra yayınlanmaz hesaplanır"),
    "overall_rank_computed": ("Genel sıralama (hesaplanan)", "int (Int64)", "türetilmiş", "rank(method='min')"),
    "disabled_rank": ("Engelli grubu sıralaması", "int (Int64)", "DisabledRank (Çnk.)", "İstanbul için yok"),
    "status": ("Durum", "str", "sonuç listesi", "FINISHED / DNS / DNF / DSQ"),
    "time_text": ("Süre metni (h:mm:ss veya mm:ss)", "str", "sonuç listesi", "Görüntüleme için ham hali korunur"),
    "swim_seconds": ("Süre - toplam saniye", "int (Int64)", "sonuç listesi", "Data-science için sayısal değer ms atılır"),
}

def build_meta(df, cfg):
    rows = []
    for c in df.columns:
        desc, dtype, source, note = COLUMN_DOCS.get(c, ("", "", "", ""))
        rows.append({
            "column": c,
            "description": desc,
            "dtype": str(df[c].dtype),
            "source": source,
            "notes": note,
        })
    notes = [
        ("NOTE", f"Veriler RaceResult Event {cfg.event_id} ({cfg.name}) resmi listelerinden alınmıştır.",
         "", f"https://{_provider_domain('my')}/{cfg.event_id}", ""),
        ("NOTE", "swim_seconds bir 'duration' (süre) olduğu için epoch zaman damgası DEĞİL toplam saniye olarak tutulur.",
         "int", "h:mm:ss -> int(seconds)", ""),
        ("NOTE", "Sıralamalar 'min' yöntemiyle hesaplanır (eşit süreler aynı sırayı paylaşır). Resmi sıralama varsa o korunur.",
         "int", "pandas rank(method='min')", ""),
        ("NOTE", "Veri çekimi racekit kütüphanesiyle yarış klasöründeki fetch_results.py üzerinden yapılır (bkz. README).",
         "text", "python fetch_results.py", ""),
    ]
    for n in notes:
        rows.append(dict(zip(("column", "description", "dtype", "source", "notes"), n)))
    return pd.DataFrame(rows)

def style_excel(path):
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            letter = col_cells[0].column_letter
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)
    wb.save(path)

def write_outputs(df, cfg):
    cfg.data_dir.mkdir(parents=True, exist_ok=True)
    stem = cfg.dataset_stem()
    xlsx_path = cfg.data_dir / f"{stem}.xlsx"
    csv_path = cfg.data_dir / f"{stem}.csv"
    meta_df = build_meta(df, cfg)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="data", index=False)
        meta_df.to_excel(writer, sheet_name="meta", index=False)
    style_excel(xlsx_path)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    return xlsx_path, csv_path

def _snapshot_path(cfg):
    data_dir = Path(cfg.data_dir)
    snap_root = data_dir / SNAPSHOT_DIR
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    snap_dir = snap_root / ts
    i = 1
    while snap_dir.exists():
        snap_dir = snap_root / f"{ts}_{i}"
        i += 1
    return snap_dir

def snapshot_previous(cfg, keep):
    data_dir = Path(cfg.data_dir)
    stem = cfg.dataset_stem()
    files = [data_dir / f"{stem}.csv", data_dir / f"{stem}.xlsx"]
    files += sorted(data_dir.glob("*_raw.json"))
    existing = [p for p in files if p.exists()]
    if not existing:
        return None

    snap_dir = _snapshot_path(cfg)
    snap_dir.mkdir(parents=True, exist_ok=True)
    for p in existing:
        shutil.copy2(p, snap_dir / p.name)

    keep_n = max(int(cfg.snapshot_keep if keep is None else keep), 1)
    snapshots = sorted(snap_dir.parent.glob("*"))
    for old in snapshots[:-keep_n]:
        if old.is_dir():
            shutil.rmtree(old, ignore_errors=True)
    return snap_dir

def list_snapshots(cfg):
    snap_root = Path(cfg.data_dir) / SNAPSHOT_DIR
    if not snap_root.exists():
        return []
    return sorted(p for p in snap_root.iterdir() if p.is_dir())

def summarize(df, cfg):
    print("\n" + "=" * 66)
    print(f"ÖZET — {cfg.name}")
    print("=" * 66)
    print(f"Toplam kayıtlı yarışmacı : {len(df)}")
    if "status" in df.columns:
        print(df["status"].value_counts().to_string())

    fin = df[df["status"] == "FINISHED"]
    if not fin.empty:
        best = fin.loc[fin["swim_seconds"].idxmin()]
        print(
            f"En hızlı genel: {best.get('full_name', '')} "
            f"({best.get('time_text', '')}) - {best.get('age_group', '')}"
        )

    if cfg.default_bib is not None:
        user = df[df["bib"].astype(str).str.strip() == str(cfg.default_bib)]
        if len(user) == 1:
            u = user.iloc[0]
            print(f"\nVarsayılan kayıt (bib {cfg.default_bib}):")
            for c in ("full_name", "gender", "age_group", "status",
                      "gender_rank", "overall_rank", "age_group_rank", "time_text", "swim_seconds"):
                print(f"   {c:24s}: {u[c]}")
        else:
            print(f"\n!! bib {cfg.default_bib} benzersiz bulunamadı ({len(user)} kayıt).")